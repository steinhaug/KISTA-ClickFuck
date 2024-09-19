from bs4 import BeautifulSoup
import time
import openpyxl
from selenium import webdriver

def read_excel_file(file_name):
    wb = openpyxl.load_workbook(file_name)
    sheet = wb.active
    leads = [cell.value for row in sheet.iter_rows() for cell in row]
    return leads


def login_to_facebook(driver, email, password):
    driver.get("https://www.facebook.com/")
    time.sleep(5)
    email_input = driver.find_element("email")
    password_input = driver.find_element("pass")
    email_input.send_keys(email)
    password_input.send_keys(password)
    login_button = driver.find_element("login")
    login_button.click()
    time.sleep(5)  # Wait for the login to complete, adjust this if needed

def scrape_contact_info(target_url, email, password):
    
    driver = webdriver.Chrome()

    // playableTile__artworkLink
    try:
        login_to_facebook(driver, email, password)

        driver.get(target_url)
        time.sleep(5)
        resp = driver.page_source
        driver.quit()

        soup = BeautifulSoup(resp, 'html.parser')

        contact_info = {"address": None, "number": None, "email": None}

        contact_items = soup.select('div[data-pagelet="ProfileContactSection"] span[dir="ltr"]')

        for contact in contact_items:
            if '@' in contact.text:
                contact_info["email"] = contact.text.strip()
            elif len(contact.text.split("-")) > 2:
                contact_info["number"] = contact.text.strip()
            elif len(contact.text.split(",")) > 2:
                contact_info["address"] = contact.text.strip()

        return contact_info

    except Exception as e:
        print(f"An error occurred: {e}")
        driver.quit()
        return None

def save_to_excel(file_name, data):
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.append(["Keyword", "Address", "Number", "Email"])

    for item in data:
        keyword = item["keyword"]
        address = item["address"]
        number = item["number"]
        email = item["email"]
        sheet.append([keyword, address, number, email])

    wb.save(file_name)

def main():
    try:
        file_name = "leadList.xlsx"  # Replace with your Excel file name
        email = "your_facebook_email_or_username"
        password = "your_facebook_password"
        keywords = read_excel_file(file_name)
        all_results = []

        for keyword in keywords:
            target_url = f"https://www.facebook.com/{keyword}"
            print(f"Searching for pages related to '{keyword}'...")
            contact_info = scrape_contact_info(target_url, email, password)
            if contact_info:
                contact_info["keyword"] = keyword
                print("Contact Information:")
                print(contact_info)
                print("-" * 50)
                all_results.append(contact_info)

        if all_results:
            save_to_excel("search_results.xlsx", all_results)
            print("Search results saved to 'search_results.xlsx'.")

    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    main()
